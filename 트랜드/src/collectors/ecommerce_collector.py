import io
import logging
import re
from datetime import datetime, timedelta

import requests
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_BRACKET_TAG_RE = re.compile(r"^\[([^\]]{2,20})\]")
_BRACKET_TAG_EXCLUDE = ("한정", "특가", "세일", "이벤트", "무료배송", "당일발송", "품절임박", "재입고")


def _extract_bracket_tag(name):
    """상품명 앞머리의 "[태그]" 표기 — "브랜드" 필드가 유통사/제조사명(예: "동국제약")인
    경우가 많은 반면, 이 태그는 실제 제품 라인/서브브랜드명인 경우가 많아(예: "[그레이온]
    웨스트 카트 브랜드 효과 3세트") 급상승 브랜드/제품 후보에 별도로 추가한다(2026-08-21)."""
    if not isinstance(name, str):
        return None
    m = _BRACKET_TAG_RE.match(name.strip())
    if not m:
        return None
    tag = m.group(1).strip()
    if re.search(r"\d", tag) or any(kw in tag for kw in _BRACKET_TAG_EXCLUDE):
        return None
    return tag

RAW_BASE = "https://raw.githubusercontent.com/primexx98-sudo/online-mall-ranking/master/data/daily"
# 2026-08-21: "카카오선물하기" 시트는 online-mall-ranking의 crawlers/kakao.py가 서브카테고리
# 2개(건강식품·영양제, 다이어트·이너뷰티)를 top_n=10씩 순차로 이어붙여 20행으로 저장한다
# (crawlers/config.py의 PLATFORMS["카카오선물하기"]["categories"] 순서 그대로, 시트 내 카테고리
# 컬럼 값은 이 두 그룹과 무관한 상품별 세부 분류라 행 위치로만 구분 가능). 기존엔 앞 10행(첫
# 서브카테고리)만 읽어 두 번째 서브카테고리가 통째로 누락됐던 걸 발견해 두 플랫폼 키로 분리.
KAKAO_SUBCATEGORIES = [
    ("카카오선물하기_건강식품", "건강식품·영양제"),
    ("카카오선물하기_다이어트", "다이어트·이너뷰티"),
]
PLATFORMS = [key for key, _ in KAKAO_SUBCATEGORIES] + ["다이소몰", "올리브영"]
TOP_N = 10


def _fetch_workbook():
    """online-mall-ranking 저장소의 최신 일별 xlsx를 오늘부터 최대 3일 역순으로 시도"""
    for delta in range(3):
        date_str = (datetime.now() - timedelta(days=delta)).strftime("%Y-%m-%d")
        url = f"{RAW_BASE}/{date_str[:7]}/{date_str}.xlsx"
        try:
            resp = requests.get(url, timeout=15)
            if resp.status_code == 200:
                logger.info(f"이커머스 판매순위 데이터 확보: {date_str}")
                return load_workbook(io.BytesIO(resp.content), data_only=True), date_str
        except Exception as e:
            logger.warning(f"이커머스 판매순위 수집 실패 [{date_str}]: {e}")
    return None, None


def get_ecommerce_rankings():
    wb, date_str = _fetch_workbook()
    if wb is None:
        logger.warning("이커머스 판매순위 데이터 없음 (최근 3일 모두 실패)")
        return {}

    # 2026-08-24: online-mall-ranking이 A열에 제품 이미지를 직접 임베드하는 '이미지' 컬럼을
    # 앞에 추가하면서(엑셀에서 좌측 썸네일로 보이도록) 기존 카테고리~이미지URL 컬럼이 전부
    # 한 칸씩 밀림. 이 컬렉터는 이미지URL '텍스트' 컬럼만 읽으면 되므로(임베드된 그림 자체는
    # 안 씀) 인덱스만 +1씩 밀어서 맞춘다.
    def _row_to_item(r):
        return {
            "rank": r[2],
            "category": r[1],
            "name": r[3],
            "brand": r[4],
            "price": r[5],
            "url": r[6],
            "image": r[7] if len(r) > 7 else "",
        }

    result = {"date": date_str}

    if "카카오선물하기" in wb.sheetnames:
        ws = wb["카카오선물하기"]
        rows = list(ws.iter_rows(min_row=2, max_row=1 + TOP_N * 2, values_only=True))
        rows = [r for r in rows if r and r[2] is not None]
        for i, (key, _label) in enumerate(KAKAO_SUBCATEGORIES):
            chunk = rows[i * TOP_N:(i + 1) * TOP_N]
            result[key] = [_row_to_item(r) for r in chunk]
    else:
        for key, _label in KAKAO_SUBCATEGORIES:
            result[key] = []

    for platform in ("다이소몰", "올리브영"):
        if platform not in wb.sheetnames:
            result[platform] = []
            continue
        ws = wb[platform]
        rows = list(ws.iter_rows(min_row=2, max_row=1 + TOP_N, values_only=True))
        result[platform] = [_row_to_item(r) for r in rows if r and r[2] is not None]

    return result


def get_rising_brand_candidates(data, limit=15):
    """당일 신규진입·순위상승(▲3 이상) 상품에서 브랜드명을 추출해 급상승 브랜드/제품
    리포트의 검색량 조회 후보로 삼는다. attach_rank_changes()로 badge가 붙은 뒤 호출해야 함.
    브랜드명 자체에 통계적 의미가 있는 건 아니고, 이커머스에서 눈에 띄게 움직인 상품의
    브랜드를 시장 신호 후보로만 사용 — 상승폭이 클수록/등장 빈도가 높을수록 가중치 부여."""
    weights = {}
    for platform in PLATFORMS:
        for it in data.get(platform, []):
            badge = it.get("badge")
            if badge == "new":
                weight = 1
            elif badge and badge.startswith("up:") and int(badge.split(":")[1]) >= 3:
                weight = 1 + int(badge.split(":")[1])
            else:
                continue
            candidates = set()
            brand = (it.get("brand") or "").strip()
            if brand:
                candidates.add(brand)
            tag = _extract_bracket_tag(it.get("name"))
            if tag:
                candidates.add(tag)
            for name in candidates:
                weights[name] = weights.get(name, 0) + weight
    ranked = sorted(weights.items(), key=lambda x: -x[1])
    return [name for name, _ in ranked[:limit]]


def attach_rank_changes(data, prev_data):
    """전일 스냅샷과 비교해 상품별 순위 변동 배지를 붙인다 (상품명+브랜드로 매칭).
    badge: "new" | "same" | "up:N" | "down:N" — prev_data 없으면 아무 배지도 안 붙임."""
    if not prev_data:
        return data
    for platform in PLATFORMS:
        prev_rank_by_key = {
            (it.get("name"), it.get("brand")): it.get("rank")
            for it in prev_data.get(platform, [])
        }
        for it in data.get(platform, []):
            key = (it.get("name"), it.get("brand"))
            prev_rank = prev_rank_by_key.get(key)
            if prev_rank is None:
                it["badge"] = "new"
            else:
                diff = prev_rank - it["rank"]
                if diff > 0:
                    it["badge"] = f"up:{diff}"
                elif diff < 0:
                    it["badge"] = f"down:{abs(diff)}"
                else:
                    it["badge"] = "same"
    return data
